import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill

# Загружаем файл, пропуская первые две строки
file_path = "ferro.xlsx"

wb_src = load_workbook(file_path)
ws_src = wb_src.active

# Значение в ячейке P3
p3_value = ws_src["P3"].value

# Заменяем на "Кол-во", если другое
if p3_value != "Кол-во":
    ws_src["P3"].value = "Кол-во"
    wb_src.save(file_path)
    print("Значение в P3 исправлено на 'Кол-во'.")

df = pd.read_excel(file_path, sheet_name="Снимок экрана", header=1, skiprows=[0])

# Группируем по коду и наименованию ТМЦ, суммируем количество
result = (
    df.groupby(["Код ТМЦ", "Наименование"], as_index=False).agg({
          "Кол-во": "sum",
          "Сумма": "sum"
      })
)

# Сохраняем результат в новый Excel
out_file = "ТМЦ_сумма.xlsx"
result.to_excel(out_file, index=False)

wb = load_workbook(out_file)
ws = wb.active

# Определяем стиль границ
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Цвет для строк с большой суммой
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # светло-красный
yellow_fill = PatternFill(start_color="ECF96A", end_color="ECF96A", fill_type="solid") # светло-желтый


# Найдём индекс колонки "Сумма"
col_sum = None
for i, cell in enumerate(ws[1], start=1):  # первая строка = заголовки
    if cell.value == "Сумма":
        col_sum = i
        break

# Применяем стили
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    row_sum = row[col_sum - 1].value  # значение из колонки "Сумма"
    for cell in row:
        cell.border = thin_border
        if col_sum and cell.column == col_sum:
            cell.number_format = '#\u00A0##0.00'
    # Если сумма > 50000 → закрашиваем всю строку
    if col_sum and isinstance(row_sum, (int, float)) and row_sum > 50000:
        for cell in row:
            cell.fill = yellow_fill
    # Если сумма > 100000 → закрашиваем всю строку
    if col_sum and isinstance(row_sum, (int, float)) and row_sum > 100000:
        for cell in row:
            cell.fill = red_fill



wb.save(out_file)
print(f"Файл сохранен с границами и форматированием: {out_file}")